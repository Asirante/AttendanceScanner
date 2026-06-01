"""GUI 통합 앱용 코어 DB 레이어."""

import os
import sqlite3
from datetime import date, datetime
from typing import List, Optional, Tuple

# 상태 전이 — 모드별 정의
# 4단계: 출근/외출/복귀/퇴근
TRANSITIONS_FULL = {
    None: ["출근"],
    "출근": ["외출", "퇴근"],
    "외출": ["복귀"],
    "복귀": ["외출", "퇴근"],
    "퇴근": ["퇴근갱신"],
}
# 2단계: 출근/퇴근만
TRANSITIONS_SIMPLE = {
    None: ["출근"],
    "출근": ["퇴근"],
    "퇴근": ["퇴근갱신"],
}
# 하위 호환용 별칭
TRANSITIONS = TRANSITIONS_FULL


class AttendanceDB:
    def __init__(self, db_path: str, mode: str = "full"):
        self.db_path = db_path
        self.mode = (
            mode  # "full"(출/외/복/퇴) 또는 "simple"(출/퇴)
        )

        parent = os.path.dirname(os.path.abspath(db_path))
        os.makedirs(parent, exist_ok=True)

        self.conn = sqlite3.connect(
            db_path, check_same_thread=False
        )
        self.conn.row_factory = sqlite3.Row
        self._init_schema()

    def set_mode(self, mode: str):
        """근태 모드 변경: 'full' 또는 'simple'."""
        self.mode = mode

    def _transitions(self):
        return (
            TRANSITIONS_SIMPLE
            if self.mode == "simple"
            else TRANSITIONS_FULL
        )

    def _init_schema(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS Employees (
                emp_id     TEXT  PRIMARY KEY,
                name       TEXT  NOT NULL,
                embedding  BLOB  NOT NULL,
                created_at TEXT  NOT NULL
            );
            CREATE TABLE IF NOT EXISTS AttendanceLog (
                log_id  INTEGER  PRIMARY KEY AUTOINCREMENT,
                emp_id  TEXT     NOT NULL,
                date    TEXT     NOT NULL,
                time    TEXT     NOT NULL,
                state   TEXT     NOT NULL,
                FOREIGN KEY (emp_id) REFERENCES Employees(emp_id)
            );
            CREATE INDEX IF NOT EXISTS idx_log_emp_date
                ON AttendanceLog(emp_id, date);
            """)
        self.conn.commit()

    def next_emp_id(self) -> str:
        """E001, E002 ... 형식으로 다음 ID 생성 (기존 최대값+1)."""
        rows = self.conn.execute(
            "SELECT emp_id FROM Employees WHERE emp_id LIKE 'E%'"
        ).fetchall()
        max_n = 0
        for r in rows:
            tail = r["emp_id"][1:]
            if tail.isdigit():
                max_n = max(max_n, int(tail))
        return f"E{max_n + 1:03d}"

    def add_employee(
        self, name: str, emb_bytes: bytes
    ) -> str:
        """이름 + 임베딩 BLOB 으로 직원 등록. 생성된 emp_id 반환."""
        emp_id = self.next_emp_id()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.conn.execute(
            "INSERT INTO Employees (emp_id, name, embedding, created_at) "
            "VALUES (?,?,?,?)",
            (emp_id, name, emb_bytes, now),
        )
        self.conn.commit()
        return emp_id

    def delete_employee(self, emp_id: str):
        """직원 + 해당 직원의 모든 근태 로그 삭제."""
        self.conn.execute(
            "DELETE FROM AttendanceLog WHERE emp_id=?",
            (emp_id,),
        )
        self.conn.execute(
            "DELETE FROM Employees WHERE emp_id=?",
            (emp_id,),
        )
        self.conn.commit()

    def list_employees(
        self, keyword: str = ""
    ) -> List[sqlite3.Row]:
        """직원 목록 조회(검색어로 이름/ID 필터)."""
        if keyword:
            like = f"%{keyword}%"
            return self.conn.execute(
                "SELECT emp_id, name, created_at FROM Employees "
                "WHERE name LIKE ? OR emp_id LIKE ? ORDER BY emp_id",
                (like, like),
            ).fetchall()
        return self.conn.execute(
            "SELECT emp_id, name, created_at FROM Employees ORDER BY emp_id"
        ).fetchall()

    def get_employee_name(self, emp_id: str) -> str:
        row = self.conn.execute(
            "SELECT name FROM Employees WHERE emp_id=?",
            (emp_id,),
        ).fetchone()
        return row["name"] if row else emp_id

    def load_all_embeddings(self) -> dict:
        """{emp_id: {"name": str, "embedding": np.ndarray(512,) float32}}"""
        import numpy as np

        rows = self.conn.execute(
            "SELECT emp_id, name, embedding FROM Employees"
        ).fetchall()
        result = {}
        for r in rows:
            emb = np.frombuffer(
                r["embedding"], dtype=np.float32
            ).copy()
            result[r["emp_id"]] = {
                "name": r["name"],
                "embedding": emb,
            }
        return result

    def get_last_state(self, emp_id: str) -> Optional[str]:
        today = date.today().isoformat()
        row = self.conn.execute(
            "SELECT state FROM AttendanceLog "
            "WHERE emp_id=? AND date=? ORDER BY log_id DESC LIMIT 1",
            (emp_id, today),
        ).fetchone()
        return row["state"] if row else None

    def get_available_actions(
        self, emp_id: str
    ) -> List[str]:
        return self._transitions().get(
            self.get_last_state(emp_id), []
        )

    def log_action(self, emp_id: str, state: str):
        valid = self.get_available_actions(emp_id)
        if state not in valid:
            raise ValueError(
                f"유효하지 않은 전이: {self.get_last_state(emp_id)} → {state}"
            )
        now = datetime.now()
        today = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        if state == "퇴근갱신":
            # 오늘의 가장 최근 '퇴근' 로그의 시각만 갱신 (새 줄 추가 안 함)
            row = self.conn.execute(
                "SELECT log_id FROM AttendanceLog "
                "WHERE emp_id=? AND date=? AND state='퇴근' "
                "ORDER BY log_id DESC LIMIT 1",
                (emp_id, today),
            ).fetchone()
            if row:
                self.conn.execute(
                    "UPDATE AttendanceLog SET time=? WHERE log_id=?",
                    (time_str, row["log_id"]),
                )
                self.conn.commit()
            return
        self.conn.execute(
            "INSERT INTO AttendanceLog (emp_id, date, time, state) VALUES (?,?,?,?)",
            (emp_id, today, time_str, state),
        )
        self.conn.commit()

    def query_logs(
        self,
        keyword: str = "",
        date_from: str = "",
        date_to: str = "",
    ) -> List[Tuple]:
        """근태 로그 조회. 이름/ID 키워드 + 날짜 범위 필터."""
        sql = (
            "SELECT a.emp_id, e.name, a.date, a.time, a.state "
            "FROM AttendanceLog a LEFT JOIN Employees e ON a.emp_id = e.emp_id "
            "WHERE 1=1"
        )
        params = []
        if keyword:
            sql += " AND (e.name LIKE ? OR a.emp_id LIKE ?)"
            like = f"%{keyword}%"
            params += [like, like]
        if date_from:
            sql += " AND a.date >= ?"
            params.append(date_from)
        if date_to:
            sql += " AND a.date <= ?"
            params.append(date_to)
        sql += " ORDER BY a.date DESC, a.time DESC"
        return [
            tuple(r)
            for r in self.conn.execute(
                sql, params
            ).fetchall()
        ]

    def export_logs_xlsx(
        self, path: str, rows: List[Tuple]
    ):
        """조회 결과를 엑셀로 저장."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "근태기록"
        headers = ["직원ID", "이름", "날짜", "시각", "상태"]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="2563EB")
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
        for r in rows:
            ws.append(list(r))

        for col, w in zip("ABCDE", [12, 16, 14, 12, 10]):
            ws.column_dimensions[col].width = w
        wb.save(path)

    def close(self):
        self.conn.close()
